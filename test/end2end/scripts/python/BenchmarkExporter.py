from Exporter import Exporter
import ModelRunner
import openpyxl
import os
import Solver
class BenchmarkExporter(Exporter):

    def __init__(self, fileName=""):
        self._fileName=fileName
        self.solvers={}
        self.workbook = openpyxl.Workbook()
        self.sheet_main = self.workbook.active
        self.sheet_stats = self.workbook.create_sheet("stats")

        self.bold_style = openpyxl.styles.NamedStyle(name="bold_style")
        self.bold_style.font = openpyxl.styles.Font(bold=True)

        self.font_red = openpyxl.styles.Font( color="FF0000")

        self.font_yellow= openpyxl.styles.Font(  color="FFFF00")

    def get_file_name(self):
        base_name, _= os.path.splitext(self._fileName)
        return base_name + '.xlsx'

    def sanifyString(self, s):
        try:
            s = s.replace("\n", " - ")
            s = s.replace(",", ";")
            return s
        except:
          return s

    def writeHeader(self, mr: ModelRunner):
        hdr = "Name,Expected_Obj,Variables,Int_Variables,Constraints,Nnz"
        for (_,r) in enumerate(mr.getRuns()):
            sname = r[-1]["solver"]
            hdr += f",{sname}-Obj,{sname}-Time,{sname}-Status"
        for (_,r) in enumerate(mr.getRuns()):
            hdr += ",{}-SolverMsg".format(r[-1]["solver"])
        header_list = hdr.split(',')
        for col_num, header_text in enumerate(header_list, 1):
            cell=self.sheet_main.cell(row=1, column=col_num, value=header_text)
            cell.style =self.bold_style
    def getModelsStats(self, run):
        if not "modelStats" in run[-1]:
            return None
        stats = run[-1]["modelStats"]
        return [ stats["nvars"], stats["nintvars"], stats["nconstr"], stats["nnz"]]
  
    def getStyle(self, r):
        try:
            if "solved" in r[-1]["timelimit"]:
                return None
            elif "limit" in r[-1]["timelimit"]:
                return self.font_yellow
            else:
                return self.font_red
        except: 
            return self.font_red

    def writeLastResultLine(self, mr: ModelRunner):
        i = len( mr.getRuns()[0] )
        m = mr._models[i-1]
        res = [m.getName(), m.getExpectedObjective()]
       
        stats = self.getModelsStats(mr.getRuns()[0])
        if stats != None:
            res.extend(stats)
        else:
            res.extend(["-","-","-","-"])  
            
        styles = [None for _ in res]

        for r in mr.getRuns():
            res.extend([
              self._getDictMemberOrMissingStr(r[-1], "objective"),
              self._getDictMemberOrMissingStr(r[-1], "solutionTime"),
              self._getDictMemberOrMissingStr(r[-1], "timelimit")])
            style=self.getStyle(r)
            styles.extend([style for _ in range(3)])

        for r in mr.getRuns():
             res.append(
                self._getDictMemberOrMissingStr(r[-1], "outmsg"))
             styles.append(self.getStyle(r))

        for col_num, header_text in enumerate(res,1):
            cell=self.sheet_main.cell(row=len( mr.getRuns()[0])+1, column=col_num, 
                                 value=header_text)
            if styles[col_num-1] is not None:
                cell.font=styles[col_num-1]
                
    def _getDictMemberOrMissingStr(self, dct, key):
        try:
            return dct[key]
        except:
            return "-"
    def addToDict(self, sname, item, value):
        self.solvers[sname][item]=self.solvers[sname][item]+value

    def collectSolverStats(self, runs):
        for r in runs:
            lastRun = r[-1]
            sname=lastRun["solver"]
            if isinstance(sname, Solver.Solver):
                sname=sname.getName()
            if "Skipped" in lastRun["outmsg"]:
                self.addToDict(sname, "failed", 1)
                return
            outcome=lastRun["timelimit"]
            self.addToDict(sname, "time_all", lastRun["solutionTime"]) 
           
            if not isinstance(r, str):
                self.addToDict(sname, "failed", 1)
            else:
                if "solved" in outcome:
                   self.addToDict(sname, "solved", 1)
                   self.addToDict(sname, "time_solved", lastRun["solutionTime"]) 
                elif "limit" in outcome:
                    self.addToDict(sname, "timelimit", 1)
                else:
                    self.addToDict(sname, "failed", 1)


    def initSolverStats(self, runs):
         for r in runs:
            lastRun = r[-1]
            sname=lastRun["solver"]
            if isinstance(sname, Solver.Solver):
                sname=sname.getName()
            self.solvers[sname]={
                "solved" : 0,
                "failed" : 0,
                "timelimit" : 0,
                "time_solved" : 0,
                "time_all" :0
            }
    def writeSolverStats(self):
        header = ["Solver"]
        header.extend(self.solvers[list(self.solvers.keys())[0]].keys())

        for col_num, header_text in enumerate(header, 1):
            cell=self.sheet_stats.cell(row=1, column=col_num, value=header_text)
            cell.style =self.bold_style 
        for row_num, (name, properties) in enumerate(self.solvers.items(), 2):
            self.sheet_stats.cell(row=row_num, column=1, value=name)
            for col_num, value in enumerate(properties.values(), 2):
                self.sheet_stats.cell(row=row_num, column=col_num, value=value)


    def exportInstanceResults(self, mr: ModelRunner):
        i = len( mr.getRuns()[0] )
        if i == 1:
            self.writeHeader(mr) 
            self.initSolverStats(mr.getRuns())
        self.writeLastResultLine(mr)
        self.collectSolverStats(mr.getRuns())
        self.writeSolverStats()
        self.workbook.save(self.get_file_name())



